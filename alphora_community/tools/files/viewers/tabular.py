"""
表格类文件查看器 - 处理 Excel、CSV、TSV 文件

核心改进：
1. 参考 excel_converter 的输出格式，显示行号和列字母坐标
2. 处理合并单元格
3. 自动显示所有 Sheet 名称
4. 智能参数推断（有 keyword 自动进入搜索模式）
"""
import os
import datetime
from typing import Optional, List, Tuple, Dict, Any
from openpyxl.utils import get_column_letter

from ..utils.common import get_file_info, clean_text


class TabularViewer:
    """表格类文件查看器"""

    SUPPORTED_EXTENSIONS = {'.xlsx', '.xls', '.csv', '.tsv'}

    def __init__(self, file_path: str):
        """
        初始化查看器

        Args:
            file_path: 文件路径
        """
        self.file_path = file_path
        self.file_info = get_file_info(file_path)
        self.ext = self.file_info['extension']

        # Excel 特有属性
        self._wb = None
        self._sheet_names: List[str] = []

    def view(
            self,
            purpose: str = "preview",
            keyword: Optional[str] = None,
            max_rows: int = 50,
            columns: Optional[str] = None,
            start_row: Optional[int] = None,
            end_row: Optional[int] = None,
            sheet_name: Optional[str] = None,
    ) -> str:
        """
        查看表格文件内容

        Args:
            purpose: 查看目的（preview/structure/search/range/stats）
            keyword: 搜索关键词
            max_rows: 最大返回行数
            columns: 要查看的列（逗号分隔）
            start_row: 起始行号（从1开始）
            end_row: 结束行号
            sheet_name: Excel 工作表名称

        Returns:
            格式化的文件内容字符串
        """
        # ============ 智能参数推断和校验 ============
        purpose, warnings = self._infer_and_validate_params(
            purpose, keyword, start_row, end_row
        )

        # CSV/TSV 处理
        if self.ext in {'.csv', '.tsv'}:
            return self._view_csv(purpose, keyword, max_rows, columns, start_row, end_row, warnings)

        # Excel 处理
        return self._view_excel(purpose, keyword, max_rows, columns, start_row, end_row, sheet_name, warnings)

    def _infer_and_validate_params(
            self,
            purpose: str,
            keyword: Optional[str],
            start_row: Optional[int],
            end_row: Optional[int]
    ) -> Tuple[str, List[str]]:
        """
        智能推断和校验参数

        核心改进：
        - 如果有 keyword 但 purpose 不是 search，自动切换并提示
        - 如果有 start_row/end_row 但 purpose 不是 range，自动切换

        Returns:
            (调整后的purpose, 警告信息列表)
        """
        warnings = []

        # 智能推断：有 keyword 应该是 search 模式
        if keyword and purpose != "search":
            warnings.append(f"⚠️ 检测到 keyword='{keyword}'，已自动切换为 search 模式")
            purpose = "search"

        # 智能推断：有 start_row/end_row 应该是 range 模式（除非在搜索）
        if (start_row is not None or end_row is not None) and purpose not in ("search", "range"):
            warnings.append(f"⚠️ 检测到行范围参数，已自动切换为 range 模式")
            purpose = "range"

        # 校验：search 模式必须有 keyword
        if purpose == "search" and not keyword:
            warnings.append("⚠️ search 模式需要提供 keyword 参数，已切换为 preview 模式")
            purpose = "preview"

        return purpose, warnings

    # ==================== CSV/TSV 处理 ====================

    def _view_csv(
            self,
            purpose: str,
            keyword: Optional[str],
            max_rows: int,
            columns: Optional[str],
            start_row: Optional[int],
            end_row: Optional[int],
            warnings: List[str]
    ) -> str:
        """处理 CSV/TSV 文件"""
        import pandas as pd

        delimiter = '\t' if self.ext == '.tsv' else ','

        try:
            df = pd.read_csv(self.file_path, delimiter=delimiter, encoding='utf-8')
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(self.file_path, delimiter=delimiter, encoding='gbk')
            except Exception as e:
                return f"❌ 读取文件失败: {e}"

        return self._process_dataframe(
            df=df,
            purpose=purpose,
            keyword=keyword,
            max_rows=max_rows,
            columns=columns,
            start_row=start_row,
            end_row=end_row,
            sheet_info=None,
            warnings=warnings
        )

    # ==================== Excel 处理 ====================

    def _view_excel(
            self,
            purpose: str,
            keyword: Optional[str],
            max_rows: int,
            columns: Optional[str],
            start_row: Optional[int],
            end_row: Optional[int],
            sheet_name: Optional[str],
            warnings: List[str]
    ) -> str:
        """处理 Excel 文件"""
        import openpyxl

        try:
            self._wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=False)
        except Exception as e:
            return f"❌ Excel 加载失败: {e}"

        self._sheet_names = self._wb.sheetnames

        # 列出所有 sheet
        if sheet_name == "__all__":
            return self._list_all_sheets()

        # ============ 全局搜索模式（跨所有 Sheet）============
        # 如果是搜索模式且没有指定 sheet，则全局搜索
        if purpose == "search" and sheet_name is None:
            return self._search_all_sheets(keyword, max_rows, warnings)

        # ============ 单 Sheet 操作 ============
        # 确定目标 sheet
        target_sheet, error = self._resolve_sheet_name(sheet_name)
        if error:
            return error

        ws = self._wb[target_sheet]

        # 处理合并单元格
        self._unmerge_and_fill(ws)

        # 获取有效数据边界
        bounds = self._get_valid_bounds(ws)
        if bounds[0] is None:
            return self._format_header(sheet_info={
                'name': target_sheet,
                'all_sheets': self._sheet_names
            }, warnings=warnings) + "\n\n(空表/无有效数据)"

        min_row, max_row, min_col, max_col = bounds

        # 根据 purpose 处理
        if purpose == "structure":
            return self._get_excel_structure(ws, target_sheet, bounds, warnings)
        elif purpose == "stats":
            return self._get_excel_stats(ws, target_sheet, bounds, warnings)
        elif purpose == "search":
            # 指定了 sheet 的搜索
            return self._search_in_excel(ws, target_sheet, bounds, keyword, max_rows, warnings)
        elif purpose == "range":
            return self._get_excel_range(ws, target_sheet, bounds, start_row, end_row, max_rows, columns, warnings)
        else:  # preview
            return self._preview_excel(ws, target_sheet, bounds, max_rows, columns, warnings)

    def _list_all_sheets(self) -> str:
        """列出所有工作表信息"""
        lines = [
            f"📊 文件: {self.file_info['name']}",
            f"📦 大小: {self.file_info['size_human']}",
            f"📋 工作表数量: {len(self._sheet_names)}",
            "",
            "【工作表列表】"
        ]

        for i, name in enumerate(self._sheet_names, 1):
            ws = self._wb[name]
            # 获取数据范围
            bounds = self._get_valid_bounds(ws)
            if bounds[0] is None:
                size_info = "(空表)"
            else:
                min_row, max_row, min_col, max_col = bounds
                rows = max_row - min_row + 1
                cols = max_col - min_col + 1
                size_info = f"{rows} 行 × {cols} 列"

            lines.append(f"  {i}. {name} - {size_info}")

        lines.append("")
        lines.append("💡 提示: 使用 sheet_name 参数指定要查看的工作表")

        return '\n'.join(lines)

    def _search_all_sheets(
            self,
            keyword: str,
            max_rows: int,
            warnings: List[str]
    ) -> str:
        """
        全局搜索：在所有 Sheet 中搜索关键词

        这是 AI Agent 最常用的搜索场景：不知道数据在哪个 sheet，
        需要工具帮助定位。

        返回格式：
        - 每个匹配结果都标注 Sheet 名称、行号、列位置
        - 按 Sheet 分组展示
        - 提供精确的定位信息供 Agent 后续操作
        """
        lines = [
            f"📊 文件: {self.file_info['name']}",
            f"📋 工作表: {len(self._sheet_names)} 个 - {self._sheet_names}",
            f"🔍 全局搜索: '{keyword}'",
            ""
        ]

        if warnings:
            for w in warnings:
                lines.append(f"# {w}")
            lines.append("")

        keyword_lower = keyword.lower()
        all_results = []  # [(sheet_name, row_num, col_letter, cell_value, row_preview)]
        sheets_with_matches = []

        # 遍历所有 Sheet 搜索
        for sheet_name in self._sheet_names:
            ws = self._wb[sheet_name]

            # 处理合并单元格
            self._unmerge_and_fill(ws)

            # 获取有效边界
            bounds = self._get_valid_bounds(ws)
            if bounds[0] is None:
                continue

            min_row, max_row, min_col, max_col = bounds
            sheet_matches = []

            # 搜索该 Sheet
            for r in range(min_row, max_row + 1):
                row_matched = False
                matched_cells = []  # 该行匹配的单元格
                row_data = {}  # 该行所有数据

                for c in range(min_col, min(max_col + 1, min_col + 30)):
                    val = ws.cell(row=r, column=c).value
                    cell_str = self._format_cell_value(val)
                    col_letter = get_column_letter(c)
                    row_data[col_letter] = cell_str

                    if keyword_lower in cell_str.lower():
                        row_matched = True
                        matched_cells.append((col_letter, cell_str))

                if row_matched:
                    # 构建行预览（显示前几列）
                    preview_cols = list(row_data.items())[:6]
                    row_preview = " | ".join([f"{k}:{v[:20]}" for k, v in preview_cols if v])

                    for col_letter, cell_value in matched_cells:
                        sheet_matches.append({
                            'row': r,
                            'col': col_letter,
                            'value': cell_value,
                            'row_preview': row_preview
                        })

            if sheet_matches:
                sheets_with_matches.append(sheet_name)
                all_results.append((sheet_name, sheet_matches))

        # 统计结果
        total_matches = sum(len(matches) for _, matches in all_results)

        if total_matches == 0:
            lines.append(f"❌ 未找到包含 '{keyword}' 的数据")
            lines.append("")
            lines.append("已搜索的工作表:")
            for sn in self._sheet_names:
                lines.append(f"  - {sn}")
            return '\n'.join(lines)

        lines.append(f"✅ 找到 {total_matches} 处匹配，分布在 {len(sheets_with_matches)} 个工作表中")
        lines.append(f"📍 匹配的工作表: {sheets_with_matches}")
        lines.append("")

        # 按 Sheet 分组输出结果
        displayed_count = 0
        for sheet_name, matches in all_results:
            if displayed_count >= max_rows:
                break

            lines.append(f"━━━ 工作表: {sheet_name} ({len(matches)} 处匹配) ━━━")

            for match in matches:
                if displayed_count >= max_rows:
                    break

                # 格式: [Sheet名!单元格位置] 匹配值 | 行预览
                cell_ref = f"{sheet_name}!{match['col']}{match['row']}"
                lines.append(f"  [{cell_ref}] \"{match['value']}\"")
                lines.append(f"      行数据: {match['row_preview']}")
                displayed_count += 1

            lines.append("")

        if total_matches > max_rows:
            lines.append(f"⚠️ 结果过多，只显示前 {max_rows} 条。可指定 sheet_name 缩小范围。")

        # 添加使用提示
        lines.append("")
        lines.append("💡 后续操作提示:")
        lines.append(f"  - 查看特定工作表: view_file(..., sheet_name='{sheets_with_matches[0]}')")
        lines.append(f"  - 查看特定行范围: view_file(..., sheet_name='...', start_row=N, end_row=M)")

        return '\n'.join(lines)

    def _resolve_sheet_name(self, sheet_name: Optional[str]) -> Tuple[str, Optional[str]]:
        """
        解析工作表名称

        Returns:
            (目标sheet名, 错误信息)
        """
        if not sheet_name:
            return self._wb.active.title, None

        if sheet_name in self._sheet_names:
            return sheet_name, None

        # 模糊匹配
        for s in self._sheet_names:
            if sheet_name.lower() in s.lower():
                return s, None

        # 未找到
        error = f"❌ 找不到工作表 '{sheet_name}'\n\n"
        error += "可用的工作表：\n"
        for i, name in enumerate(self._sheet_names, 1):
            error += f"  {i}. {name}\n"
        return "", error

    def _unmerge_and_fill(self, ws):
        """拆解所有合并单元格，并将左上角的值填充到整个区域"""
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col = merged_range.min_col
            min_row = merged_range.min_row
            max_col = merged_range.max_col
            max_row = merged_range.max_row
            top_left_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(
                start_row=min_row, start_column=min_col,
                end_row=max_row, end_column=max_col
            )
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = top_left_value

    def _get_valid_bounds(self, ws) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
        """获取有效数据边界（跳过空行空列）"""
        min_row = ws.max_row + 1
        max_row = 0
        min_col = ws.max_column + 1
        max_col = 0
        found = False

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    found = True
                    min_row = min(min_row, cell.row)
                    max_row = max(max_row, cell.row)
                    min_col = min(min_col, cell.column)
                    max_col = max(max_col, cell.column)

        if not found:
            return None, None, None, None
        return min_row, max_row, min_col, max_col

    def _format_header(
            self,
            sheet_info: Optional[Dict[str, Any]] = None,
            warnings: Optional[List[str]] = None
    ) -> str:
        """格式化文件头信息"""
        lines = [f"📊 文件: {self.file_info['name']}"]

        if sheet_info:
            lines.append(f"📋 当前工作表: {sheet_info['name']}")
            all_sheets = sheet_info.get('all_sheets', [])
            if len(all_sheets) > 1:
                other_sheets = [s for s in all_sheets if s != sheet_info['name']]
                lines.append(f"📑 其他工作表: {', '.join(other_sheets)}")

        if warnings:
            lines.append("")
            for w in warnings:
                lines.append(w)

        return '\n'.join(lines)

    def _format_cell_value(self, val) -> str:
        """格式化单元格值"""
        if val is None:
            return ""
        if isinstance(val, datetime.datetime):
            if val.hour == 0 and val.minute == 0 and val.second == 0:
                return val.strftime("%Y-%m-%d")
            return val.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(val, datetime.time):
            return val.strftime("%H:%M:%S")
        if isinstance(val, datetime.date):
            return val.strftime("%Y-%m-%d")
        val_str = str(val).strip()
        val_str = val_str.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
        val_str = " ".join(val_str.split())
        # 限制单元格内容长度
        if len(val_str) > 50:
            val_str = val_str[:47] + "..."
        return val_str

    def _preview_excel(
            self,
            ws,
            sheet_name: str,
            bounds: Tuple[int, int, int, int],
            max_rows: int,
            columns: Optional[str],
            warnings: List[str]
    ) -> str:
        """
        预览 Excel 内容（参考 excel_converter 格式）

        输出格式示例（单 sheet）：
        # All Sheets: ['Sheet1']
        # Inspecting Sheet: Sheet1 | Range: A1:F20
        Idx,A,B,C,D
        1,日期,销售额,备注,类型
        2,2024-01-01,100,测试,A

        输出格式示例（多 sheet，不展示数据行）：
        # All Sheets: ['Sheet1', 'Data', 'Summary']
        # Sheet Count: 3
        # Sheet Details:
        #   1. Sheet1 - 100 行 × 5 列 (A1:E100)
        #   2. Data - 50 行 × 3 列 (A1:C50)
        #   3. Summary - 10 行 × 2 列 (A1:B10)
        # Tip: 使用 sheet_name 参数指定要查看的工作表
        """
        min_row, max_row, min_col, max_col = bounds

        # 多 sheet 情况：只展示概要信息，不展示数据行
        if len(self._sheet_names) > 1:
            lines = [
                f"# All Sheets: {self._sheet_names}",
                f"# Sheet Count: {len(self._sheet_names)}",
                f"# Sheet Details:"
            ]

            for i, name in enumerate(self._sheet_names, 1):
                sheet_ws = self._wb[name]
                sheet_bounds = self._get_valid_bounds(sheet_ws)
                if sheet_bounds[0] is None:
                    size_info = "(空表)"
                else:
                    s_min_row, s_max_row, s_min_col, s_max_col = sheet_bounds
                    rows = s_max_row - s_min_row + 1
                    cols = s_max_col - s_min_col + 1
                    range_str = f"{get_column_letter(s_min_col)}{s_min_row}:{get_column_letter(s_max_col)}{s_max_row}"
                    size_info = f"{rows} 行 × {cols} 列 ({range_str})"
                lines.append(f"#   {i}. {name} - {size_info}")

            # 添加警告
            if warnings:
                for w in warnings:
                    lines.append(f"# {w}")

            lines.append(f"#")
            lines.append(f"# 💡 Tip: 使用 sheet_name 参数指定要查看的工作表数据")

            return '\n'.join(lines)

        # 单 sheet 情况：正常展示数据行
        # 限制最大列数
        max_scan_col = min(max_col, min_col + 29)  # 最多30列

        # 构建头部
        lines = [
            f"# All Sheets: {self._sheet_names}",
            f"# Inspecting Sheet: {sheet_name} | Range: {get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        ]

        # 添加警告
        if warnings:
            for w in warnings:
                lines.append(f"# {w}")

        total_rows = max_row - min_row + 1
        if total_rows > max_rows:
            lines.append(f"# Warning: Data truncated. Showing first {max_rows} rows out of {total_rows}.")

        # 构建列头（Idx + 列字母）
        header_row = ["Idx"]
        for c in range(min_col, max_scan_col + 1):
            header_row.append(get_column_letter(c))
        lines.append(",".join(header_row))

        # 构建数据行
        process_rows = min(max_row - min_row + 1, max_rows)
        for r in range(min_row, min_row + process_rows):
            row_data = [str(r)]
            for c in range(min_col, max_scan_col + 1):
                cell_val = ws.cell(row=r, column=c).value
                clean_val = self._format_cell_value(cell_val)
                # 处理 CSV 特殊字符
                if ',' in clean_val or '"' in clean_val:
                    clean_val = '"' + clean_val.replace('"', '""') + '"'
                row_data.append(clean_val)
            lines.append(",".join(row_data))

        return '\n'.join(lines)

    def _get_excel_structure(
            self,
            ws,
            sheet_name: str,
            bounds: Tuple[int, int, int, int],
            warnings: List[str]
    ) -> str:
        """获取 Excel 结构信息"""
        import pandas as pd

        min_row, max_row, min_col, max_col = bounds
        total_rows = max_row - min_row + 1
        total_cols = max_col - min_col + 1

        lines = [
            f"# All Sheets: {self._sheet_names}",
            f"# Inspecting Sheet: {sheet_name}",
            f"# Data Range: {get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
            f"# Size: {total_rows} 行 × {total_cols} 列",
            ""
        ]

        if warnings:
            for w in warnings:
                lines.append(f"# {w}")
            lines.append("")

        lines.append("【列信息】")

        # 获取表头行（假设第一行是表头）
        header_row = min_row
        for c in range(min_col, min(max_col + 1, min_col + 30)):
            col_letter = get_column_letter(c)
            header_val = ws.cell(row=header_row, column=c).value
            header_name = self._format_cell_value(header_val) or f"(列{col_letter})"

            # 采样分析列数据
            sample_values = []
            non_null_count = 0
            for r in range(min_row + 1, min(max_row + 1, min_row + 101)):  # 采样100行
                val = ws.cell(row=r, column=c).value
                if val is not None and str(val).strip():
                    non_null_count += 1
                    if len(sample_values) < 3:
                        sample_values.append(self._format_cell_value(val)[:20])

            sample_str = ", ".join(sample_values) if sample_values else "(无数据)"
            lines.append(f"  {col_letter}. {header_name}")
            lines.append(f"     非空: {non_null_count}/100 | 示例: {sample_str}")

        return '\n'.join(lines)

    def _get_excel_stats(
            self,
            ws,
            sheet_name: str,
            bounds: Tuple[int, int, int, int],
            warnings: List[str]
    ) -> str:
        """获取 Excel 统计信息"""
        import pandas as pd

        # 读取数据到 DataFrame 进行统计
        min_row, max_row, min_col, max_col = bounds

        data = []
        headers = []

        # 获取表头
        for c in range(min_col, max_col + 1):
            val = ws.cell(row=min_row, column=c).value
            headers.append(self._format_cell_value(val) or f"Col{c}")

        # 获取数据
        for r in range(min_row + 1, max_row + 1):
            row_data = []
            for c in range(min_col, max_col + 1):
                val = ws.cell(row=r, column=c).value
                row_data.append(val)
            data.append(row_data)

        df = pd.DataFrame(data, columns=headers)

        lines = [
            f"# All Sheets: {self._sheet_names}",
            f"# Inspecting Sheet: {sheet_name}",
            f"# Size: {len(df)} 行 × {len(df.columns)} 列",
            ""
        ]

        if warnings:
            for w in warnings:
                lines.append(f"# {w}")
            lines.append("")

        # 数值列统计
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        if numeric_cols:
            lines.append("【数值列统计】")
            stats_df = df[numeric_cols].describe().round(2)
            lines.append(stats_df.to_string())
            lines.append("")

        # 非数值列概览
        non_numeric_cols = [c for c in df.columns if c not in numeric_cols]
        if non_numeric_cols:
            lines.append("【非数值列概览】")
            for col in non_numeric_cols[:10]:
                unique_count = df[col].nunique()
                top_values = df[col].value_counts().head(3)
                top_str = ', '.join([f"{v}({c})" for v, c in top_values.items()])
                lines.append(f"  - {col}: {unique_count}个唯一值")
                if top_str:
                    lines.append(f"    最常见: {top_str}")

        return '\n'.join(lines)

    def _search_in_excel(
            self,
            ws,
            sheet_name: str,
            bounds: Tuple[int, int, int, int],
            keyword: str,
            max_rows: int,
            warnings: List[str]
    ) -> str:
        """在 Excel 中搜索关键词"""
        min_row, max_row, min_col, max_col = bounds

        lines = [
            f"# All Sheets: {self._sheet_names}",
            f"# Inspecting Sheet: {sheet_name}",
            f"# Search: '{keyword}'",
            ""
        ]

        if warnings:
            for w in warnings:
                lines.append(f"# {w}")
            lines.append("")

        # 搜索匹配的行
        matched_rows = []
        keyword_lower = keyword.lower()

        for r in range(min_row, max_row + 1):
            row_matched = False
            row_data = [str(r)]

            for c in range(min_col, min(max_col + 1, min_col + 30)):
                val = ws.cell(row=r, column=c).value
                cell_str = self._format_cell_value(val)

                if keyword_lower in cell_str.lower():
                    row_matched = True
                    # 高亮标记匹配
                    cell_str = f"**{cell_str}**"

                if ',' in cell_str or '"' in cell_str:
                    cell_str = '"' + cell_str.replace('"', '""') + '"'
                row_data.append(cell_str)

            if row_matched:
                matched_rows.append(row_data)

        if not matched_rows:
            lines.append(f"未找到包含 '{keyword}' 的数据")
            return '\n'.join(lines)

        lines.append(f"# Found: {len(matched_rows)} 条匹配记录")
        if len(matched_rows) > max_rows:
            lines.append(f"# Warning: 只显示前 {max_rows} 条")
        lines.append("")

        # 构建列头
        header_row = ["Idx"]
        for c in range(min_col, min(max_col + 1, min_col + 30)):
            header_row.append(get_column_letter(c))
        lines.append(",".join(header_row))

        # 输出匹配行
        for row_data in matched_rows[:max_rows]:
            lines.append(",".join(row_data))

        return '\n'.join(lines)

    def _get_excel_range(
            self,
            ws,
            sheet_name: str,
            bounds: Tuple[int, int, int, int],
            start_row: Optional[int],
            end_row: Optional[int],
            max_rows: int,
            columns: Optional[str],
            warnings: List[str]
    ) -> str:
        """获取指定范围的 Excel 数据"""
        min_row, max_row, min_col, max_col = bounds

        # 计算实际范围
        if end_row is not None and end_row < 0:
            # 负数表示最后 N 行
            actual_start = max(min_row, max_row + end_row + 1)
            actual_end = max_row
        elif start_row is not None:
            actual_start = max(min_row, start_row)
            if end_row is not None:
                actual_end = min(max_row, end_row)
            else:
                actual_end = min(max_row, actual_start + max_rows - 1)
        else:
            actual_start = min_row
            actual_end = min(max_row, min_row + max_rows - 1)

        total_rows = max_row - min_row + 1

        lines = [
            f"# All Sheets: {self._sheet_names}",
            f"# Inspecting Sheet: {sheet_name}",
            f"# Range: Row {actual_start} to {actual_end} (Total: {total_rows} rows)",
            ""
        ]

        if warnings:
            for w in warnings:
                lines.append(f"# {w}")
            lines.append("")

        # 构建列头
        max_scan_col = min(max_col, min_col + 29)
        header_row = ["Idx"]
        for c in range(min_col, max_scan_col + 1):
            header_row.append(get_column_letter(c))
        lines.append(",".join(header_row))

        # 构建数据行
        for r in range(actual_start, actual_end + 1):
            row_data = [str(r)]
            for c in range(min_col, max_scan_col + 1):
                cell_val = ws.cell(row=r, column=c).value
                clean_val = self._format_cell_value(cell_val)
                if ',' in clean_val or '"' in clean_val:
                    clean_val = '"' + clean_val.replace('"', '""') + '"'
                row_data.append(clean_val)
            lines.append(",".join(row_data))

        return '\n'.join(lines)

    # ==================== DataFrame 通用处理 ====================

    def _process_dataframe(
            self,
            df,
            purpose: str,
            keyword: Optional[str],
            max_rows: int,
            columns: Optional[str],
            start_row: Optional[int],
            end_row: Optional[int],
            sheet_info: Optional[Dict[str, Any]],
            warnings: List[str]
    ) -> str:
        """处理 DataFrame（用于 CSV/TSV）"""
        import pandas as pd

        total_rows, total_cols = df.shape
        col_names = df.columns.tolist()

        # 筛选列
        if columns:
            selected_cols = [c.strip() for c in columns.split(',')]
            valid_cols = []
            for col in selected_cols:
                if col in col_names:
                    valid_cols.append(col)
                else:
                    for c in col_names:
                        if col.lower() in c.lower():
                            valid_cols.append(c)
                            break
            if valid_cols:
                df = df[valid_cols]
            else:
                return f"❌ 指定的列不存在。\n\n可用的列：{', '.join(col_names)}"

        # 根据 purpose 处理
        if purpose == "structure":
            return self._format_csv_structure(df, total_rows, warnings)
        elif purpose == "stats":
            return self._format_csv_stats(df, warnings)
        elif purpose == "search":
            return self._search_in_csv(df, keyword, max_rows, warnings)
        elif purpose == "range":
            return self._format_csv_range(df, total_rows, start_row, end_row, max_rows, warnings)
        else:  # preview
            return self._format_csv_preview(df, total_rows, max_rows, warnings)

    def _format_csv_preview(
            self,
            df,
            total_rows: int,
            max_rows: int,
            warnings: List[str]
    ) -> str:
        """预览 CSV 内容"""
        lines = [
            f"# File: {self.file_info['name']}",
            f"# Size: {total_rows} rows × {len(df.columns)} columns",
        ]

        if warnings:
            for w in warnings:
                lines.append(f"# {w}")

        if total_rows > max_rows:
            lines.append(f"# Warning: Showing first {max_rows} rows")

        lines.append("")

        # 使用 DataFrame 的 to_csv 但不带 index
        preview_df = df.head(max_rows)
        csv_content = preview_df.to_csv(index=True, index_label='Idx')
        lines.append(csv_content.strip())

        return '\n'.join(lines)

    def _format_csv_structure(self, df, total_rows: int, warnings: List[str]) -> str:
        """获取 CSV 结构信息"""
        lines = [
            f"# File: {self.file_info['name']}",
            f"# Size: {total_rows} rows × {len(df.columns)} columns",
            ""
        ]

        if warnings:
            for w in warnings:
                lines.append(f"# {w}")
            lines.append("")

        lines.append("【列信息】")
        for i, col in enumerate(df.columns, 1):
            dtype = str(df[col].dtype)
            non_null = df[col].notna().sum()
            sample_values = df[col].dropna().head(3).tolist()
            sample_str = ', '.join([str(v)[:20] for v in sample_values])

            lines.append(f"  {i}. {col}")
            lines.append(f"     类型: {dtype} | 非空: {non_null}/{len(df)}")
            if sample_str:
                lines.append(f"     示例: {sample_str}")

        return '\n'.join(lines)

    def _format_csv_stats(self, df, warnings: List[str]) -> str:
        """获取 CSV 统计信息"""
        lines = [
            f"# File: {self.file_info['name']}",
            f"# Size: {len(df)} rows × {len(df.columns)} columns",
            ""
        ]

        if warnings:
            for w in warnings:
                lines.append(f"# {w}")
            lines.append("")

        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        if numeric_cols:
            lines.append("【数值列统计】")
            stats_df = df[numeric_cols].describe().round(2)
            lines.append(stats_df.to_string())

        return '\n'.join(lines)

    def _search_in_csv(
            self,
            df,
            keyword: str,
            max_rows: int,
            warnings: List[str]
    ) -> str:
        """在 CSV 中搜索"""
        mask = df.astype(str).apply(
            lambda x: x.str.contains(keyword, case=False, na=False)
        ).any(axis=1)
        matched_df = df[mask]

        lines = [
            f"# File: {self.file_info['name']}",
            f"# Search: '{keyword}'",
            f"# Found: {len(matched_df)} matches",
            ""
        ]

        if warnings:
            for w in warnings:
                lines.append(f"# {w}")
            lines.append("")

        if len(matched_df) == 0:
            lines.append(f"未找到包含 '{keyword}' 的数据")
            return '\n'.join(lines)

        if len(matched_df) > max_rows:
            lines.append(f"# Warning: Showing first {max_rows} matches")
            lines.append("")

        display_df = matched_df.head(max_rows)
        lines.append(display_df.to_csv(index=True, index_label='Idx').strip())

        return '\n'.join(lines)

    def _format_csv_range(
            self,
            df,
            total_rows: int,
            start_row: Optional[int],
            end_row: Optional[int],
            max_rows: int,
            warnings: List[str]
    ) -> str:
        """获取指定范围的 CSV 数据"""
        if end_row is not None and end_row < 0:
            display_df = df.tail(abs(end_row))
            actual_start = total_rows + end_row + 1
            actual_end = total_rows
        elif start_row is not None:
            start_idx = max(0, start_row - 1)
            if end_row is not None:
                end_idx = min(total_rows, end_row)
            else:
                end_idx = min(total_rows, start_idx + max_rows)
            display_df = df.iloc[start_idx:end_idx]
            actual_start = start_row
            actual_end = end_idx
        else:
            display_df = df.head(max_rows)
            actual_start = 1
            actual_end = min(max_rows, total_rows)

        lines = [
            f"# File: {self.file_info['name']}",
            f"# Range: Row {actual_start} to {actual_end} (Total: {total_rows} rows)",
            ""
        ]

        if warnings:
            for w in warnings:
                lines.append(f"# {w}")
            lines.append("")

        lines.append(display_df.to_csv(index=True, index_label='Idx').strip())

        return '\n'.join(lines)