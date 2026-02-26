"""
Robust tabular inspector for iterative data exploration.

Supported formats: CSV/TSV/TXT, Excel, JSON, Parquet.

Usage examples:
  python inspect_file.py /mnt/workspace/data.xlsx --purpose preview
  python inspect_file.py /mnt/workspace/data.xlsx --sheet __all__
  python inspect_file.py /mnt/workspace/data.xlsx --purpose search --keyword 北京
  python inspect_file.py /mnt/workspace/data.xlsx --purpose locate --keyword 销售额,订单号
  python inspect_file.py /mnt/workspace/data.csv --purpose stats
  python inspect_file.py /mnt/workspace/data.csv --purpose range --start-row 120 --end-row 160
"""

import argparse
import json
import os
import sys
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter


DEFAULT_LOCATE_KEYWORDS = [
    "id", "key", "code", "no",
    "订单", "单号", "编号", "客户", "用户",
    "日期", "时间", "金额", "销售", "收入", "成本", "利润",
    "区域", "城市", "省", "产品", "sku", "渠道", "状态",
]


def detect_encoding(filepath: str) -> str:
    """Detect text encoding with safe fallback."""
    try:
        import chardet

        with open(filepath, "rb") as f:
            raw = f.read(min(os.path.getsize(filepath), 100_000))
        result = chardet.detect(raw)
        return result.get("encoding") or "utf-8"
    except ImportError:
        return "utf-8"


def detect_separator(filepath: str, encoding: str) -> str:
    candidates = [",", "\t", ";", "|"]
    with open(filepath, "r", encoding=encoding, errors="replace") as f:
        sample = "".join(f.readline() for _ in range(10))
    counts = {sep: sample.count(sep) for sep in candidates}
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ","


def _safe_df(path: str, encoding: str, sheet=None):
    ext = Path(path).suffix.lower()

    if ext in (".csv", ".tsv", ".txt"):
        sep = "\t" if ext == ".tsv" else detect_separator(path, encoding)
        for enc in [encoding, "utf-8", "gbk", "gb18030", "latin1", "cp1252"]:
            try:
                return pd.read_csv(path, sep=sep, encoding=enc, low_memory=False), {
                    "encoding": enc,
                    "separator": sep,
                }
            except Exception:
                continue
        raise ValueError("Cannot decode CSV/TSV with known encodings.")

    if ext in (".xlsx", ".xls"):
        return pd.read_excel(path, sheet_name=sheet if sheet is not None else 0, engine="openpyxl"), {
            "encoding": None,
            "separator": None,
        }

    if ext == ".json":
        with open(path, "r", encoding=encoding, errors="replace") as f:
            data = json.load(f)
        if isinstance(data, list):
            return pd.DataFrame(data), {"encoding": encoding, "separator": None}
        if isinstance(data, dict):
            if any(isinstance(v, list) for v in data.values()):
                return pd.DataFrame(data), {"encoding": encoding, "separator": None}
            return pd.DataFrame([data]), {"encoding": encoding, "separator": None}
        raise ValueError(f"Unsupported JSON root type: {type(data).__name__}")

    if ext == ".parquet":
        return pd.read_parquet(path), {"encoding": None, "separator": None}

    return pd.read_csv(path, encoding=encoding, low_memory=False), {
        "encoding": encoding,
        "separator": ",",
    }


def _sheet_names(path: str):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _render_excel_inventory(path: str):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    print("# Excel Inventory (compact)")
    print(f"# Sheet count: {len(names)}")
    print("SheetNo,SheetName,Rows,Cols,HeaderPreview")

    for idx, sn in enumerate(names, start=1):
        ws = wb[sn]
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        headers = []
        max_header_cols = min(cols, 8)
        for c in range(1, max_header_cols + 1):
            headers.append(_fmt_value(ws.cell(row=1, column=c).value))
        header_preview = " | ".join([h for h in headers if h]) or "(empty)"
        header_preview = header_preview.replace('"', '""')
        print(f'{idx},{sn},{rows},{cols},"{header_preview}"')

    wb.close()
    print("# Next step: choose one sheet and inspect progressively with --sheet <name> --rows 5")


def _search_all_sheets(filepath: str, encoding: str, keyword: str, max_rows: int):
    names = _sheet_names(filepath)
    print(f"# All Sheets: {names}")
    print(f"# Search: '{keyword}'")
    print("# Mode: global across all sheets")

    keyword_lower = keyword.lower()
    total_matches = 0
    shown = 0
    print("HitNo,Sheet,RowRef,ColRef,Header,Value,RowPreview")

    for sn in names:
        try:
            df, _ = _safe_df(filepath, encoding, sheet=sn)
        except Exception:
            continue

        cols = list(df.columns)
        for row_no, (_, row) in enumerate(df.iterrows(), start=1):
            row_preview = " | ".join([f"{c}={_fmt_value(row[c])}" for c in cols[:6]])
            for col_idx, col in enumerate(cols, start=1):
                cell = _fmt_value(row[col])
                if keyword_lower in cell.lower():
                    total_matches += 1
                    if shown < max_rows:
                        shown += 1
                        value = cell.replace('"', '""')
                        preview = row_preview.replace('"', '""')
                        print(
                            f'{shown},{sn},{sn}!R{row_no},{get_column_letter(col_idx)},{col},"{value}","{preview}"'
                        )

    print(f"# Found: {total_matches}")
    if total_matches > max_rows:
        print(f"# Warning: 只显示前 {max_rows} 条。可使用 --sheet 限定到单个工作表。")


def _fmt_value(v):
    if pd.isna(v):
        return ""
    text = str(v).replace("\n", " ").replace("\r", " ").replace("\t", " ").strip()
    text = " ".join(text.split())
    if len(text) > 80:
        return text[:77] + "..."
    return text


def _field_ref(col_idx_1based: int, col_name: str) -> str:
    return f"{get_column_letter(col_idx_1based)}:{col_name}"


def _sheet_label(sheet_name: str = None) -> str:
    return sheet_name if sheet_name else "DATA"


def _parse_keywords(keyword: str):
    if keyword:
        words = [w.strip() for w in keyword.split(",") if w.strip()]
        if words:
            return words
    return DEFAULT_LOCATE_KEYWORDS


def _infer_and_validate_params(purpose: str, keyword: str, start_row: int, end_row: int):
    warnings = []
    final_purpose = purpose

    if keyword and final_purpose != "search":
        warnings.append(f"⚠️ 检测到 keyword='{keyword}'，已自动切换为 search 模式")
        final_purpose = "search"

    if (start_row is not None or end_row is not None) and final_purpose not in ("search", "range"):
        warnings.append("⚠️ 检测到行范围参数，已自动切换为 range 模式")
        final_purpose = "range"

    if final_purpose == "search" and not keyword:
        warnings.append("⚠️ search 模式需要提供 keyword 参数，已切换为 preview 模式")
        final_purpose = "preview"

    return final_purpose, warnings


def _print_warnings(warnings):
    for w in warnings:
        print(f"# {w}")


def _print_field_map(df: pd.DataFrame):
    print("# Field map:")
    for i, col in enumerate(df.columns, start=1):
        print(f"#   {get_column_letter(i)} -> {col}")


def _print_locator_grid(df: pd.DataFrame, rows: int, sheet_name: str, start_row_no: int = 1):
    display = df.head(rows)
    header = ["Idx", "RowRef"] + [_field_ref(i, str(c)) for i, c in enumerate(display.columns, start=1)]
    print(",".join(header))
    for i, (_, row) in enumerate(display.iterrows(), start=0):
        row_no = start_row_no + i
        row_ref = f"{_sheet_label(sheet_name)}!R{row_no}"
        vals = [str(row_no), row_ref]
        for val in row.tolist():
            text = _fmt_value(val)
            if "," in text or '"' in text:
                text = '"' + text.replace('"', '""') + '"'
            vals.append(text)
        print(",".join(vals))


def _print_basic_header(filepath: str, ftype: str, meta: dict):
    size = os.path.getsize(filepath)
    print(f"# File: {os.path.basename(filepath)}")
    print(f"# Path: {filepath}")
    print(f"# Type: {ftype}")
    print(f"# Size: {size:,} bytes ({size / 1024 / 1024:.2f} MB)")
    if meta.get("encoding"):
        print(f"# Encoding: {meta['encoding']}")
    if meta.get("separator"):
        print(f"# Separator: '{meta['separator']}'")


def _apply_columns(df: pd.DataFrame, columns: str):
    if not columns:
        return df
    wanted = [c.strip() for c in columns.split(",") if c.strip()]
    if not wanted:
        return df

    resolved = []
    for w in wanted:
        if w in df.columns:
            resolved.append(w)
            continue
        lw = w.lower()
        for c in df.columns:
            if lw == str(c).lower() or lw in str(c).lower():
                resolved.append(c)
                break
    resolved = list(dict.fromkeys(resolved))
    if not resolved:
        raise ValueError(f"Columns not found: {wanted}. Available: {list(df.columns)}")
    return df[resolved]


def render_structure(df: pd.DataFrame, sheet_name: str = None):
    print(f"# Shape: {len(df):,} rows x {len(df.columns)} columns")
    print(f"# Sheet: {_sheet_label(sheet_name)}")
    _print_field_map(df)
    print()
    print("## Columns")
    for col in df.columns:
        s = df[col]
        null_count = int(s.isna().sum())
        null_pct = (null_count / len(df) * 100) if len(df) else 0.0
        unique = int(s.nunique(dropna=True))
        sample = [_fmt_value(x) for x in s.dropna().head(3).tolist()]
        sample_str = "; ".join([x for x in sample if x]) or "(no sample)"
        print(
            f"- {col}: dtype={s.dtype}, non_null={len(df)-null_count}/{len(df)}, "
            f"null={null_pct:.1f}%, unique={unique}, sample={sample_str}"
        )


def render_preview(df: pd.DataFrame, rows: int, sheet_name: str = None):
    n = min(rows, len(df))
    print(f"# Preview: first {n} rows")
    print(f"# Sheet: {_sheet_label(sheet_name)}")
    _print_field_map(df)
    if len(df) > rows:
        print(f"# Truncated: showing {rows}/{len(df)} rows")
    _print_locator_grid(df, rows, sheet_name, start_row_no=1)


def render_stats(df: pd.DataFrame, sheet_name: str = None):
    print(f"# Shape: {len(df):,} rows x {len(df.columns)} columns")
    print(f"# Sheet: {_sheet_label(sheet_name)}")
    _print_field_map(df)
    print(f"# Duplicate rows: {int(df.duplicated().sum()):,}")
    print()

    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    if numeric_cols:
        print("## Numeric Summary")
        print(df[numeric_cols].describe().round(4).to_string())
        print()

    print("## Missing Values")
    miss = df.isna().sum().sort_values(ascending=False)
    if miss.sum() == 0:
        print("(none)")
    else:
        for col, cnt in miss.items():
            if cnt > 0:
                print(f"- {col}: {int(cnt)} ({cnt / len(df) * 100:.1f}%)")

    object_cols = [c for c in df.columns if c not in numeric_cols]
    if object_cols:
        print()
        print("## Categorical Top Values")
        for col in object_cols[:20]:
            vc = df[col].astype(str).value_counts(dropna=True).head(5)
            top = ", ".join([f"{k}({v})" for k, v in vc.items()])
            print(f"- {col}: {top or '(empty)'}")


def render_search(df: pd.DataFrame, keyword: str, rows: int, sheet_name: str = None):
    if not keyword:
        raise ValueError("keyword is required when purpose=search")
    print(f"# Search keyword: {keyword}")
    print(f"# Sheet: {_sheet_label(sheet_name)}")
    _print_field_map(df)

    keyword_lower = keyword.lower()
    hits = []
    for row_pos, (_, row) in enumerate(df.iterrows(), start=1):
        row_preview = " | ".join([f"{c}={_fmt_value(row[c])}" for c in df.columns[:6]])
        for col_pos, col_name in enumerate(df.columns, start=1):
            cell_text = _fmt_value(row[col_name])
            if keyword_lower in cell_text.lower():
                hits.append(
                    {
                        "row_no": row_pos,
                        "row_ref": f"{_sheet_label(sheet_name)}!R{row_pos}",
                        "col_ref": get_column_letter(col_pos),
                        "header": str(col_name),
                        "value": cell_text,
                        "row_preview": row_preview,
                    }
                )

    print(f"# Matches: {len(hits)}")
    if len(hits) == 0:
        return
    if len(hits) > rows:
        print(f"# Truncated: showing first {rows} matches")

    print("MatchNo,RowRef,ColRef,Header,Value,RowPreview")
    for i, h in enumerate(hits[:rows], start=1):
        value = h["value"].replace('"', '""')
        preview = h["row_preview"].replace('"', '""')
        print(
            f'{i},{h["row_ref"]},{h["col_ref"]},{h["header"]},"{value}","{preview}"'
        )


def render_locate(df: pd.DataFrame, keyword: str, rows: int, sheet_name: str = None):
    kws = _parse_keywords(keyword)
    print(f"# Locate mode")
    print(f"# Sheet: {_sheet_label(sheet_name)}")
    print(f"# Rows: {len(df):,}, Cols: {len(df.columns)}")
    print(f"# Keywords: {kws}")
    _print_field_map(df)
    print()

    # 1) Primary-key-like candidates
    key_name_tokens = ("id", "key", "code", "no", "编号", "单号", "订单")
    key_candidates = []
    for i, col in enumerate(df.columns, start=1):
        s = df[col]
        non_null = max(1, int(s.notna().sum()))
        unique = int(s.nunique(dropna=True))
        unique_ratio = unique / non_null
        non_null_ratio = non_null / max(1, len(df))
        name_hit = any(tok in str(col).lower() for tok in key_name_tokens)
        score = unique_ratio + 0.3 * non_null_ratio + (0.25 if name_hit else 0)
        key_candidates.append((score, col, i, unique_ratio, non_null_ratio, name_hit))

    key_candidates.sort(key=lambda x: x[0], reverse=True)
    print("## Key Candidates (top)")
    print("Rank,FieldRef,UniqueRatio,NonNullRatio,NameHint")
    for rank, item in enumerate(key_candidates[:8], start=1):
        _, col, idx, u_ratio, nn_ratio, name_hit = item
        print(f"{rank},{_field_ref(idx, str(col))},{u_ratio:.3f},{nn_ratio:.3f},{name_hit}")
    print()

    # 2) Business-field candidates by header/value keyword hits
    sample_df = df.head(min(len(df), 20_000))
    biz_scores = []
    lower_kws = [k.lower() for k in kws]
    for i, col in enumerate(sample_df.columns, start=1):
        col_lower = str(col).lower()
        header_hits = sum(1 for k in lower_kws if k in col_lower)
        try:
            col_text = sample_df[col].astype(str)
            value_hits = 0
            for k in kws:
                value_hits += int(col_text.str.contains(k, case=False, na=False).sum())
        except Exception:
            value_hits = 0
        score = header_hits * 200 + value_hits
        biz_scores.append((score, col, i, header_hits, value_hits))

    biz_scores.sort(key=lambda x: x[0], reverse=True)
    print("## Business Field Candidates (top)")
    print("Rank,FieldRef,HeaderHitCount,ValueHitCount")
    shown = 0
    for score, col, idx, h_hit, v_hit in biz_scores:
        if shown >= 10:
            break
        if score <= 0 and shown >= 5:
            break
        shown += 1
        print(f"{shown},{_field_ref(idx, str(col))},{h_hit},{v_hit}")
    print()

    # 3) Compact hit locations
    scan_df = df.head(min(len(df), 5_000))
    hit_rows = []
    for row_no, (_, row) in enumerate(scan_df.iterrows(), start=1):
        row_kws = set()
        col_hits = []
        for col_idx, col in enumerate(scan_df.columns, start=1):
            cell = _fmt_value(row[col])
            if not cell:
                continue
            matched = [k for k in kws if k.lower() in cell.lower()]
            if matched:
                row_kws.update(matched)
                col_hits.append(f"{get_column_letter(col_idx)}:{col}")
        if col_hits:
            preview = " | ".join([f"{c}={_fmt_value(row[c])}" for c in scan_df.columns[:6]])
            hit_rows.append(
                {
                    "row_ref": f"{_sheet_label(sheet_name)}!R{row_no}",
                    "keywords": ";".join(sorted(row_kws)),
                    "fields": ";".join(col_hits[:6]),
                    "preview": preview,
                }
            )
        if len(hit_rows) >= rows:
            break

    print("## Keyword Hit Locations")
    if not hit_rows:
        print("(no hit rows)")
        return

    print("HitNo,RowRef,MatchedKeywords,MatchedFields,RowPreview")
    for i, h in enumerate(hit_rows, start=1):
        preview = h["preview"].replace('"', '""')
        print(
            f'{i},{h["row_ref"]},"{h["keywords"]}","{h["fields"]}","{preview}"'
        )


def render_range(df: pd.DataFrame, start_row: int, end_row: int, rows: int, sheet_name: str = None):
    if start_row is None and end_row is None:
        chunk = df.head(rows)
        print(f"# Range: default first {len(chunk)} rows")
        start_no = 1
    elif end_row is not None and end_row < 0:
        chunk = df.tail(abs(end_row))
        print(f"# Range: last {len(chunk)} rows")
        start_no = max(1, len(df) - len(chunk) + 1)
    else:
        s = 1 if start_row is None else max(1, start_row)
        e = min(len(df), s + rows - 1) if end_row is None else min(len(df), end_row)
        if e < s:
            raise ValueError("Invalid range: end_row is smaller than start_row.")
        chunk = df.iloc[s - 1 : e]
        print(f"# Range: rows {s}..{e}")
        start_no = s

    print(f"# Sheet: {_sheet_label(sheet_name)}")
    _print_field_map(df)
    _print_locator_grid(chunk, len(chunk), sheet_name, start_row_no=start_no)


def run_for_dataframe(df: pd.DataFrame, args, sheet_name: str = None):
    if args.columns:
        df = _apply_columns(df, args.columns)

    if args.purpose == "structure":
        render_structure(df, sheet_name=sheet_name)
    elif args.purpose == "stats":
        render_stats(df, sheet_name=sheet_name)
    elif args.purpose == "search":
        render_search(df, args.keyword, args.rows, sheet_name=sheet_name)
    elif args.purpose == "locate":
        render_locate(df, args.keyword, args.rows, sheet_name=sheet_name)
    elif args.purpose == "range":
        render_range(df, args.start_row, args.end_row, args.rows, sheet_name=sheet_name)
    else:
        render_preview(df, args.rows, sheet_name=sheet_name)


def main():
    parser = argparse.ArgumentParser(description="Robust inspector for tabular files")
    parser.add_argument("filepath", help="Path to data file")
    parser.add_argument(
        "--purpose",
        default="preview",
        choices=["preview", "structure", "stats", "search", "locate", "range"],
        help="Inspection mode",
    )
    parser.add_argument("--rows", "--max-lines", type=int, default=8, help="Max rows to display (default: 8)")
    parser.add_argument("--sheet", "--sheet-name", type=str, default=None, help="Excel sheet name/index, or __all__")
    parser.add_argument("--keyword", type=str, default=None, help="Keyword for search mode")
    parser.add_argument("--columns", type=str, default=None, help="Comma separated columns")
    parser.add_argument("--start-row", "--start_row", type=int, default=None, help="Start row (1-indexed)")
    parser.add_argument("--end-row", "--end_row", type=int, default=None, help="End row; negative means last N rows")
    parser.add_argument("--encoding", type=str, default=None, help="Force encoding for text files")
    args = parser.parse_args()

    args.purpose, warnings = _infer_and_validate_params(
        args.purpose, args.keyword, args.start_row, args.end_row
    )

    if not os.path.exists(args.filepath):
        print(f"[ERROR] File not found: {args.filepath}")
        sys.exit(1)

    ext = Path(args.filepath).suffix.lower()
    encoding = args.encoding or detect_encoding(args.filepath)

    try:
        if ext in (".xlsx", ".xls"):
            if args.sheet == "__all__":
                _print_basic_header(args.filepath, "excel", {"encoding": None, "separator": None})
                # Hard guardrail: __all__ always returns compact inventory only.
                _render_excel_inventory(args.filepath)
                return

            if args.purpose == "search" and args.sheet is None:
                _print_basic_header(args.filepath, "excel", {"encoding": None, "separator": None})
                _print_warnings(warnings)
                _search_all_sheets(args.filepath, encoding, args.keyword, args.rows)
                return

            sheet = args.sheet
            if sheet is not None:
                try:
                    sheet = int(sheet)
                except ValueError:
                    pass
            df, meta = _safe_df(args.filepath, encoding, sheet=sheet)
            _print_basic_header(args.filepath, "excel", meta)
            _print_warnings(warnings)
            print(f"# Sheet: {sheet if sheet is not None else 0}")
            sheet_name = str(sheet) if sheet is not None else "0"
            if isinstance(sheet, int):
                try:
                    sheet_name = _sheet_names(args.filepath)[sheet]
                except Exception:
                    sheet_name = str(sheet)
            run_for_dataframe(df, args, sheet_name=sheet_name)
            return

        df, meta = _safe_df(args.filepath, encoding)
        ftype = "parquet" if ext == ".parquet" else ("json" if ext == ".json" else "tabular")
        _print_basic_header(args.filepath, ftype, meta)
        _print_warnings(warnings)
        run_for_dataframe(df, args, sheet_name="DATA")
    except Exception as exc:
        print(f"[ERROR] Inspection failed: {exc}")
        import traceback

        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
