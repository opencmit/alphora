"""
Excel / CSV Reader

依赖:
- .xlsx / .xls → openpyxl (可选依赖，未安装时给出清晰提示)
- .csv → stdlib csv 模块，无额外依赖
"""

import csv
import io
import os
from datetime import datetime
from typing import Optional, Union, List, Dict, Tuple, Any

try:
    from alphora.sandbox.tools.inspector.readers import FileContent
except ImportError:
    from dataclasses import dataclass, field

    @dataclass
    class FileContent:
        text: str
        total_lines: int
        file_type: str
        size: int
        metadata: dict = field(default_factory=dict)

_MAX_COL_WIDTH = 28
_MAX_OUTPUT_CHARS = 10000
_MAX_DISPLAY_COLS = 15
_MAX_SCAN_ROWS = 500

_EXCEL_EPOCH = datetime(1899, 12, 30)


# ── value formatting ──────────────────────────────────────────────

def _try_excel_date(value: int) -> Optional[str]:
    """Convert Excel serial date number to readable date if it looks like one."""
    if not (40000 <= value <= 55000):
        return None
    try:
        from datetime import timedelta
        dt = _EXCEL_EPOCH + timedelta(days=value)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, int):
        date_str = _try_excel_date(value)
        if date_str:
            return date_str
        return str(value)
    if isinstance(value, float):
        if value == int(value) and abs(value) < 1e15:
            int_val = int(value)
            date_str = _try_excel_date(int_val)
            if date_str:
                return date_str
            return str(int_val)
        abs_v = abs(value)
        if abs_v >= 1e8:
            return f"{value:.0f}"
        if abs_v >= 100:
            return f"{value:.2f}"
        if abs_v >= 1:
            return f"{value:.4f}"
        if abs_v >= 0.0001:
            return f"{value:.6f}"
        return f"{value:.8f}"
    s = str(value)
    return s.replace("\n", " ").replace("\r", "").strip()


def _is_numeric(s: str) -> bool:
    s = s.strip()
    if not s or s in ("-", "—", "N/A", "n/a"):
        return False
    try:
        float(s.replace(",", ""))
        return True
    except ValueError:
        return False


# ── merged cell handling ──────────────────────────────────────────

def _build_merge_map(ws) -> Dict[Tuple[int, int], Any]:
    """Map (row, col) → top-left value for every cell inside a merged range."""
    merge_map: Dict[Tuple[int, int], Any] = {}
    for mr in ws.merged_cells.ranges:
        val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if r != mr.min_row or c != mr.min_col:
                    merge_map[(r, c)] = val
    return merge_map


def _get_cell(ws, row: int, col: int, merge_map: Dict) -> Any:
    if (row, col) in merge_map:
        return merge_map[(row, col)]
    return ws.cell(row, col).value


def _build_suppress_set(ws, max_col: int) -> set:
    """Cells in full-width merges where the value should NOT be repeated for display."""
    suppress = set()
    threshold = max(max_col * 0.6, 3)
    for mr in ws.merged_cells.ranges:
        col_span = mr.max_col - mr.min_col + 1
        if col_span >= threshold:
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    if r != mr.min_row or c != mr.min_col:
                        suppress.add((r, c))
    return suppress


# ── data bounds detection ─────────────────────────────────────────

def _find_bounds(ws, merge_map: Dict) -> Tuple[int, int]:
    """Find actual (max_row, max_col), ignoring trailing empties and inflated dimensions."""
    raw_mr = ws.max_row or 0
    raw_mc = ws.max_column or 0
    if raw_mr == 0:
        return 0, 0

    max_col = 0
    for mr_range in ws.merged_cells.ranges:
        max_col = max(max_col, mr_range.max_col)

    scan_cols = min(raw_mc, 120)
    for r in range(1, min(raw_mr + 1, 40)):
        for c in range(scan_cols, 0, -1):
            val = _get_cell(ws, r, c, merge_map)
            if val is not None and str(val).strip():
                max_col = max(max_col, c)
                break

    if max_col == 0:
        max_col = min(raw_mc, 20)

    max_row = 0
    check_cols = list(range(1, min(max_col + 1, 8)))
    for r in range(raw_mr, 0, -1):
        for c in check_cols:
            val = _get_cell(ws, r, c, merge_map)
            if val is not None and str(val).strip():
                max_row = r
                break
        if max_row > 0:
            break

    return max_row, max_col


# ── header detection ──────────────────────────────────────────────

def _detect_header_rows(rows: List[List[str]]) -> int:
    """Return the number of header rows (title + column headers) before data begins."""
    if len(rows) < 2:
        return 0

    for i in range(min(len(rows), 10)):
        non_empty = [c for c in rows[i] if c.strip()]
        if not non_empty:
            continue
        unique_vals = set(non_empty)
        if len(unique_vals) == 1 and len(non_empty) > 2:
            continue
        nums = sum(1 for c in non_empty if _is_numeric(c))
        texts = len(non_empty) - nums
        if nums >= 3 and nums > texts:
            return i

    return min(4, len(rows))


# ── table formatting ──────────────────────────────────────────────

def _compute_widths(rows: List[List[str]], num_cols: int) -> List[int]:
    widths = [0] * num_cols
    for r in rows:
        for i in range(min(len(r), num_cols)):
            widths[i] = max(widths[i], len(r[i]))
    return [min(w, _MAX_COL_WIDTH) for w in widths]


def _truncate_cols(
    rows: List[List[str]], max_cols: int = _MAX_DISPLAY_COLS,
) -> Tuple[List[List[str]], int, int]:
    """If columns exceed max_cols, keep left + right sides with '…' separator.

    Returns (new_rows, original_col_count, omitted_col_count).
    """
    if not rows:
        return rows, 0, 0
    orig = max(len(r) for r in rows)
    if orig <= max_cols:
        return rows, orig, 0
    left = max_cols - 3
    right = 2
    omitted = orig - left - right
    result = []
    for r in rows:
        padded = r + [""] * (orig - len(r))
        result.append(padded[:left] + ["…"] + padded[-right:])
    return result, orig, omitted


def _render_rows(rows: List[List[str]], widths: List[int], sep_after: int = -1) -> str:
    def _fr(cells):
        parts = []
        for i, c in enumerate(cells):
            w = widths[i] if i < len(widths) else _MAX_COL_WIDTH
            v = c if len(c) <= w else c[: w - 1] + "…"
            parts.append(v.ljust(w))
        return "| " + " | ".join(parts) + " |"

    sep = "|" + "|".join("-" * (w + 2) for w in widths) + "|"
    lines: List[str] = []
    for i, r in enumerate(rows):
        padded = r + [""] * (len(widths) - len(r))
        lines.append(_fr(padded[: len(widths)]))
        if i == sep_after:
            lines.append(sep)
    return "\n".join(lines)


def _fmt_table(
    all_rows: List[List[str]],
    sep_after: int = -1,
    max_chars: int = _MAX_OUTPUT_CHARS,
    orig_rows: int = 0,
    orig_cols: int = 0,
) -> str:
    """Format rows as markdown table, dynamically truncating rows/cols to fit max_chars."""
    if not all_rows:
        return "(empty)"

    all_rows, real_cols, cols_omitted = _truncate_cols(all_rows)
    if orig_cols <= 0:
        orig_cols = real_cols
    total = len(all_rows)
    if orig_rows <= 0:
        orig_rows = total

    num_cols = max(len(r) for r in all_rows)
    widths = _compute_widths(all_rows, num_cols)

    text = _render_rows(all_rows, widths, sep_after)
    shown_rows = total

    if len(text) > max_chars:
        avg = len(text) / total if total else 1
        affordable = max(int((max_chars - 200) / avg), 4)

        header_n = max(sep_after + 1, 0)
        data_budget = max(affordable - header_n, 2)
        tail_n = max(min(3, data_budget // 4), 1)
        head_n = header_n + data_budget - tail_n

        if head_n + tail_n < total:
            shown_rows = head_n + tail_n
            omitted_r = total - shown_rows
            text = _render_rows(all_rows[:head_n], widths, sep_after)
            text += f"\n  ... (省略 {omitted_r} 行) ...\n"
            text += _render_rows(all_rows[-tail_n:], widths)

    hints = []
    if shown_rows < orig_rows:
        hints.append(f"行: 显示 {shown_rows}/{orig_rows}")
    if cols_omitted > 0:
        shown_c = orig_cols - cols_omitted
        hints.append(f"列: 显示 {shown_c}/{orig_cols}")
    if hints:
        text += f"\n\n[表格截断] {', '.join(hints)}。可用 search 按关键词定位具体内容"

    if len(text) > max_chars * 1.5:
        text = text[:max_chars]
        text += f"\n\n[输出截断] 已达字符上限 (原始 {orig_rows} 行 × {orig_cols} 列)"

    return text


# ── CSV ───────────────────────────────────────────────────────────

def _read_csv(text: str, path: str, size: int, max_chars: int = _MAX_OUTPUT_CHARS) -> FileContent:
    reader = csv.reader(io.StringIO(text))
    rows: List[List[str]] = []
    total_row_count = 0
    for row in reader:
        total_row_count += 1
        if len(rows) < _MAX_SCAN_ROWS:
            rows.append(row)

    if not rows:
        return FileContent(
            text="(empty CSV file)",
            total_lines=0,
            file_type="csv",
            size=size,
            metadata={"rows": 0, "columns": 0},
        )

    str_rows = [[str(c) for c in r] for r in rows]
    num_cols = max(len(r) for r in str_rows)
    header_n = _detect_header_rows(str_rows)
    table = _fmt_table(
        str_rows,
        sep_after=header_n - 1 if header_n > 0 else -1,
        max_chars=max_chars,
        orig_rows=total_row_count,
        orig_cols=num_cols,
    )

    return FileContent(
        text=table,
        total_lines=total_row_count,
        file_type="csv",
        size=size,
        metadata={
            "rows": total_row_count,
            "columns": num_cols,
            "headers": rows[0] if rows else [],
        },
    )


# ── XLSX — overview ──────────────────────────────────────────────

def _build_overview(wb, path: str, size: int) -> FileContent:
    """Build workbook overview listing all sheets with key stats."""
    names = wb.sheetnames
    bn = os.path.basename(path)
    lines = [f"{bn} — {len(names)} sheets", ""]

    col_w_name = max(len(n) for n in names) + 2
    col_w_name = min(col_w_name, 52)
    hdr = f"  {'#':>3}  {'Sheet':<{col_w_name}}  {'Rows':>5}  {'Cols':>5}  {'Merged':>6}"
    lines.append(hdr)
    lines.append("  " + "-" * (len(hdr) - 2))

    for i, name in enumerate(names):
        ws = wb[name]
        r = ws.max_row or 0
        c = ws.max_column or 0
        m = len(list(ws.merged_cells.ranges))
        c_display = str(c) if c <= 50 else f"~{c}"
        display_name = name if len(name) <= col_w_name else name[: col_w_name - 1] + "…"
        lines.append(
            f"  {i + 1:>3}. {display_name:<{col_w_name}}  {r:>5}  {c_display:>5}  {m:>6}"
        )

    lines.append("")
    lines.append('Use sheet="<sheet name>" to inspect a specific sheet.')

    text = "\n".join(lines)
    return FileContent(
        text=text,
        total_lines=len(lines),
        file_type="excel",
        size=size,
        metadata={
            "sheets": names,
            "sheet_count": len(names),
            "active_sheet": wb.active.title,
        },
    )


# ── XLSX — single sheet ─────────────────────────────────────────

def _read_single_sheet(wb, ws, path: str, size: int, max_chars: int = _MAX_OUTPUT_CHARS) -> FileContent:
    names = wb.sheetnames
    sn = ws.title
    bn = os.path.basename(path)

    merge_map = _build_merge_map(ws)
    max_r, max_c = _find_bounds(ws, merge_map)

    if max_r == 0 or max_c == 0:
        return FileContent(
            text=f"{bn} — Sheet: {sn} (empty)",
            total_lines=0,
            file_type="excel",
            size=size,
            metadata={"sheets": names, "active_sheet": sn, "rows": 0, "columns": 0},
        )

    suppress = _build_suppress_set(ws, max_c)
    scan_r = min(max_r, _MAX_SCAN_ROWS)

    rows: List[List[str]] = []
    for r in range(1, scan_r + 1):
        row: List[str] = []
        for c in range(1, max_c + 1):
            if (r, c) in suppress:
                row.append("")
            else:
                row.append(_format_value(_get_cell(ws, r, c, merge_map)))
        rows.append(row)

    header_n = _detect_header_rows(rows)
    merged_count = len(list(ws.merged_cells.ranges))

    summary = f"{bn} — Sheet: {sn} ({max_r} rows × {max_c} cols"
    if merged_count:
        summary += f", {merged_count} merged regions"
    summary += ")"
    if len(names) > 1:
        summary += f"  [sheet {names.index(sn) + 1}/{len(names)}]"

    sep = header_n - 1 if header_n > 0 else -1
    table_text = _fmt_table(
        rows, sep_after=sep, max_chars=max_chars,
        orig_rows=max_r, orig_cols=max_c,
    )

    text = f"{summary}\n\n{table_text}"
    return FileContent(
        text=text,
        total_lines=max_r,
        file_type="excel",
        size=size,
        metadata={
            "sheets": names,
            "active_sheet": sn,
            "rows": max_r,
            "columns": max_c,
            "merged_cell_ranges": merged_count,
            "header_rows": header_n,
        },
    )


# ── XLSX entry ────────────────────────────────────────────────────

def _read_xlsx(
    data: bytes, path: str, size: int, sheet: Optional[str] = None,
    max_chars: int = _MAX_OUTPUT_CHARS,
) -> FileContent:
    try:
        import openpyxl
    except ImportError:
        return FileContent(
            text="[Error] openpyxl is required to read Excel files. Install: pip install openpyxl",
            total_lines=0,
            file_type="excel",
            size=size,
            metadata={"error": "missing_dependency"},
        )

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    try:
        if sheet is None and len(wb.sheetnames) > 1:
            return _build_overview(wb, path, size)

        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        return _read_single_sheet(wb, ws, path, size, max_chars)
    finally:
        wb.close()


# ── public entry ──────────────────────────────────────────────────

def read(
    data: Union[str, bytes],
    path: str,
    size: int,
    sheet: Optional[str] = None,
    max_chars: int = _MAX_OUTPUT_CHARS,
    **_kwargs,
) -> FileContent:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return _read_csv(
            data if isinstance(data, str) else data.decode("utf-8"), path, size, max_chars
        )
    return _read_xlsx(
        data if isinstance(data, bytes) else data.encode("utf-8"), path, size, sheet, max_chars
    )


if __name__ == "__main__":
    import re

    EXCEL_FILE = "/Users/tiantiantian/工作/1-梧桐AlphaData/3-测试数据/其他数据/测试用例/chatexcel/表格测试-交通费.xlsx"
    with open(EXCEL_FILE, "rb") as f:
        data = f.read()
    size = len(data)

    # ── 1. 多 sheet 概览 ──
    print("=" * 70)
    print("1. read(data, path, size) — 多 sheet 概览")
    print("=" * 70)
    fc = read(data, EXCEL_FILE, size)
    print(fc.text)

    # ── 2. 默认预算(10000字符)下的正常输出 ──
    print("\n" + "=" * 70)
    print('2. read(..., sheet="分省财务-收入和利润") — 默认预算')
    print("=" * 70)
    fc = read(data, EXCEL_FILE, size, sheet="分省财务-收入和利润")
    print(fc.text)
    print(f"\n--- 输出字符数: {len(fc.text)} ---")

    # ── 3. 小预算强制触发行截断 ──
    print("\n" + "=" * 70)
    print('3. read(..., sheet="分省财务-收入和利润", max_chars=3000) — 行截断')
    print("=" * 70)
    fc = read(data, EXCEL_FILE, size, sheet="分省财务-收入和利润", max_chars=3000)
    print(fc.text)
    print(f"\n--- 输出字符数: {len(fc.text)} ---")

    # ── 4. 极小预算 ──
    print("\n" + "=" * 70)
    print('4. read(..., sheet="主要经营情况表", max_chars=1500) — 极小预算')
    print("=" * 70)
    fc = read(data, EXCEL_FILE, size, sheet="主要经营情况表", max_chars=1500)
    print(fc.text)
    print(f"\n--- 输出字符数: {len(fc.text)} ---")

    # ── 5. 宽表测试（列截断）──
    print("\n" + "=" * 70)
    print('5. read(..., sheet="主要产品业务情况表") — 38列宽表,列截断')
    print("=" * 70)
    fc = read(data, EXCEL_FILE, size, sheet="主要产品业务情况表")
    print(fc.text)
    print(f"\n--- 输出字符数: {len(fc.text)} ---")

    # ── 6. 维度修正 + 截断 ──
    print("\n" + "=" * 70)
    print('6. read(..., sheet="分省业务-有线宽带业务") — 维度修正')
    print("=" * 70)
    fc = read(data, EXCEL_FILE, size, sheet="分省业务-有线宽带业务")
    print(fc.text)
    print(f"\n--- 输出字符数: {len(fc.text)} ---")

    # ── 7. 关键词搜索 ──
    print("\n" + "=" * 70)
    print('7. 关键词搜索: "云"')
    print("=" * 70)
    fc = read(data, EXCEL_FILE, size, sheet="主要经营情况表")
    keyword = "云"
    for i, line in enumerate(fc.text.splitlines(), 1):
        if keyword in line:
            print(f"  行{i}: {line.strip()}")

    # ── 8. 正则搜索 ──
    print("\n" + "=" * 70)
    print(r'8. 正则搜索: \d{3,4}\.\d{2}')
    print("=" * 70)
    fc = read(data, EXCEL_FILE, size, sheet="分省财务-收入和利润")
    pat = re.compile(r"\d{3,4}\.\d{2}")
    for i, line in enumerate(fc.text.splitlines(), 1):
        if pat.search(line):
            print(f"  行{i}: {line.strip()}")
